# coding: utf-8
from openpyxl import load_workbook
from common.setting import Settings


class DealExcel:
    def __init__(self):
        self.ws = None

    s = Settings()

    # 读取测试数据
    def read(self, sheetname=None):
        path = self.s.excel_path
        wb = load_workbook(path)
        if sheetname:
            sheet = wb[sheetname]
        else:
            sheet = wb.active
        row = sheet.max_row
        column = self.s.excel_data_column
        data_list = []
        for r in range(2, row + 1):
            value = sheet.cell(r, column).value
            if value:
                data_list.append([eval(value)])
        wb.close()
        return data_list

    # 写入数据
    def write(self, data, row_index, column_index, sheetname=None):
        path = self.s.excel_path
        wb = load_workbook(path)
        if not sheetname:
            ws = wb.active
        else:
            ws = wb[sheetname]
        # 获取Excel中原测试数据
        temp = eval(ws.cell(row_index, column_index).value)
        # 循环替换新测试数据
        for i in temp.keys():
            if i in data.keys():
                temp[i] = data[i]
            else:
                continue
        ws.cell(row_index, column_index).value = str(temp)
        wb.save(path)
        wb.close()
        return row_index
